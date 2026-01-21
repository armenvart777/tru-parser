"""
Сборка финального XLSX с результатами парсинга + валидации.

Входные данные:
- tru_data.csv - результаты парсера (матчинг + данные из каталога)
- ru_validation.json - результаты валидации РУ через Росздрав
- roszdrav_extra.json - дополнительные РУ найденные через API (опционально)

Логика:
1. МЗ РФ / ФС (старый формат) - убираем
2. Несовпадение категории (mismatch) - убираем РУ + ОКПД2 + PDF
3. Не найден в реестре (not_found) - убираем РУ + PDF
4. Подтверждён (ok) - оставляем
5. Дополняем из roszdrav_extra.json (если есть)

Запуск: python3 build_final.py
"""

import csv
import re
import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

INPUT_CSV = "tru_data.csv"
VALIDATION_JSON = "ru_validation.json"
ROSZDRAV_EXTRA = "roszdrav_extra.json"
OUTPUT_XLSX = "результат_ТРУ_финал.xlsx"


def main():
    # Загрузить CSV
    with open(INPUT_CSV, encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))

    # Загрузить валидацию
    validation = {}
    if os.path.exists(VALIDATION_JSON):
        with open(VALIDATION_JSON, encoding="utf-8") as f:
            validation = json.load(f)
    else:
        print(f"ВНИМАНИЕ: {VALIDATION_JSON} не найден")

    # Загрузить доп. РУ из Росздрав
    extra_ru = {}
    if os.path.exists(ROSZDRAV_EXTRA):
        with open(ROSZDRAV_EXTRA, encoding="utf-8") as f:
            extra_ru = json.load(f)
        print(f"Загружено доп. РУ: {len(extra_ru)}")

    stats = {
        "total": len(rows),
        "kept": 0,
        "cleared_old": 0,
        "cleared_mismatch": 0,
        "cleared_notfound": 0,
        "added_from_roszdrav": 0,
    }

    for row in rows:
        ru = row.get("ru_number", "").strip()
        row_num = row.get("row_num", "")

        if not ru:
            # Нет РУ - проверяем есть ли в доп. данных
            if row_num in extra_ru:
                extra = extra_ru[row_num]
                row["ru_number"] = extra.get("noRu", "")
                row["_status"] = "added_roszdrav"
                stats["added_from_roszdrav"] += 1
            continue

        # 1. МЗ РФ / ФС - убираем
        if re.match(r"^(МЗ\s*РФ|ФС\s)", ru):
            row["ru_number"] = ""
            row["ru_pdf_url"] = ""
            row["_status"] = "cleared_old"
            stats["cleared_old"] += 1
            continue

        # 2. Проверяем по валидации
        val = validation.get(row_num, {})
        if not val:
            # Нет данных валидации - оставляем как есть (если валидация не запускалась)
            row["_status"] = "no_validation"
            stats["kept"] += 1
            continue

        if val.get("ok"):
            row["_status"] = "validated_ok"
            stats["kept"] += 1
        else:
            reason = val.get("reason", "unknown")
            row["ru_number"] = ""
            row["ru_pdf_url"] = ""
            if reason in ("mismatch", "old_format") or "категория" in reason or "наш:" in reason:
                row["_status"] = "cleared_mismatch"
                stats["cleared_mismatch"] += 1
            else:
                row["_status"] = "cleared_notfound"
                stats["cleared_notfound"] += 1

    # Собираем XLSX
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Результат ТРУ"

    headers = ["№ строки", "Товар", "Код ТСР (полный)", "Код ТСР", "ОКПД2", "Номер РУ", "Ссылка на скан РУ"]
    ws.append(headers)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row in rows:
        # Маппинг полей
        tsr_full = row.get("matched_code", "")
        tsr_short = row.get("tsr_code", "")

        # Если нет tsr_short но есть tsr_full - извлекаем
        if tsr_full and not tsr_short:
            m = re.match(r'(\d{2}-\d{2}-\d{2})', tsr_full)
            if m:
                tsr_short = m.group(1)

        data = [
            row.get("row_num", ""),
            row.get("name", ""),
            tsr_full,
            tsr_short,
            row.get("okpd2", ""),
            row.get("ru_number", ""),
            row.get("ru_pdf_url", ""),
        ]
        ws.append(data)

        for cell in ws[ws.max_row]:
            cell.border = thin_border

    # Ширина колонок
    col_widths = [10, 65, 22, 15, 18, 22, 55]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(OUTPUT_XLSX)

    # Итоговая статистика
    final_with_ru = sum(1 for r in rows if r.get("ru_number", "").strip())
    final_with_okpd = sum(1 for r in rows if r.get("okpd2", "").strip())
    final_with_tsr = sum(1 for r in rows if r.get("matched_code", "").strip())
    final_with_pdf = sum(1 for r in rows if r.get("ru_pdf_url", "").strip())

    print(f"\nГотово: {OUTPUT_XLSX}")
    print(f"\nОчищено:")
    print(f"  МЗ РФ / ФС (старый формат): {stats['cleared_old']}")
    print(f"  Несовпадение категории:      {stats['cleared_mismatch']}")
    print(f"  Не найден в реестре:          {stats['cleared_notfound']}")
    print(f"  Подтверждено:                 {stats['kept']}")
    print(f"  Добавлено из Росздрав:        {stats['added_from_roszdrav']}")
    print(f"\nФинальные цифры:")
    print(f"  Всего строк:     {stats['total']}")
    print(f"  Код ТСР:         {final_with_tsr}")
    print(f"  ОКПД2:           {final_with_okpd}")
    print(f"  Номер РУ:        {final_with_ru}")
    print(f"  Ссылка на PDF:   {final_with_pdf}")


if __name__ == "__main__":
    main()
